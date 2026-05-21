import argparse
import dataclasses
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from vpf.tournaments import parse_list_tab, ParseWarning, Tournament
from vpf.venues import extract_schedule_hyperlinks, extract_event_hyperlinks, load_venues, slug_for_venue_display, merge_pdf_urls
from vpf.writers import write_tournaments_json, write_venues_json, write_warnings_json


def build_feed(xlsx_path: Path, venues_yml: Path, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    wb_links = load_workbook(xlsx_path)             # keep hyperlinks
    wb_data = load_workbook(xlsx_path, data_only=True)

    tournaments, warnings = parse_list_tab(wb_data["List"])
    discovered_urls = extract_schedule_hyperlinks(wb_links["Schedule 2026"])
    venues = load_venues(venues_yml)
    venues = merge_pdf_urls(venues, discovered_urls)

    # Build the venue-display -> slug lookup FIRST so we can normalize both sides
    # of the per-event hyperlink join through slugs. (The List tab calls the WSOP
    # venue "WSOP", but the Schedule tab header is "WSOP Paris/Horseshoe" — and
    # similar for Aria, MGM, Golden Nugget. Without slug normalisation those
    # venues miss the join entirely.)
    venue_slug_lookup: dict[str, str] = {}
    for t in tournaments:
        if t.venue_display in venue_slug_lookup:
            continue
        slug = slug_for_venue_display(t.venue_display, venues)
        if slug:
            venue_slug_lookup[t.venue_display] = slug
        else:
            warnings.append(ParseWarning(
                row_number=0,
                issue=f"unknown venue: {t.venue_display!r}",
                raw_row=(t.venue_display,),
            ))

    # Extract per-event hyperlinks, re-keying from (schedule-venue-display, date, event)
    # to (slug, date, event) using the same slug resolver.
    raw_event_links = extract_event_hyperlinks(wb_links["Schedule 2026"])
    event_links_by_slug: dict[tuple[str, "object", str], str] = {}
    for (sched_venue, day, ev_name), url in raw_event_links.items():
        slug = slug_for_venue_display(sched_venue, venues)
        if not slug:
            continue
        event_links_by_slug[(slug, day, ev_name.strip())] = url

    # Some venues (notably WSOP) link every cell to a single wrong PDF in the
    # source sheet. venues.yml can set override_per_event_url=true to skip the
    # per-event join and fall back to the venue-level structure_pdf_url.
    override_slugs = {v["slug"] for v in venues if v.get("override_per_event_url")}

    # Attach per-event PDF URL to each tournament via the slug-keyed lookup.
    enriched: list[Tournament] = []
    for t in tournaments:
        slug = venue_slug_lookup.get(t.venue_display)
        url: str | None = None
        if slug and slug not in override_slugs:
            url = event_links_by_slug.get((slug, t.date_pt, t.event_name.strip()))
        if url is not None:
            enriched.append(dataclasses.replace(t, structure_pdf_url=url))
        else:
            enriched.append(t)
    tournaments = enriched

    # Wizardofviz enrichment: stack size, level length, handed, rake. Best-effort —
    # any fetch/parse failure leaves these fields nil.
    from vpf.wov import fetch_events as fetch_wov_events
    wov_events = fetch_wov_events()
    wov_index = _build_wov_index(wov_events, venues)
    tournaments = [_enrich_with_wov(t, venue_slug_lookup, wov_index) for t in tournaments]

    now = datetime.now(timezone.utc).replace(tzinfo=None)  # plain UTC for ISO 'Z'

    source_mtime = datetime.fromtimestamp(xlsx_path.stat().st_mtime, tz=timezone.utc).replace(tzinfo=None)
    write_tournaments_json(
        out_dir / "tournaments.json",
        tournaments=tournaments,
        venue_slug_lookup=venue_slug_lookup,
        generated_at=now,
        source_sheet_updated_at=source_mtime,
    )
    write_venues_json(out_dir / "venues.json", venues=venues, discovered_urls=discovered_urls)
    write_warnings_json(out_dir / "parse_warnings.json", warnings=warnings, generated_at=now)


def _normalize_event_name(name: str) -> str:
    """Strip buy-in dollar prefix, expand/contract game abbreviations, lowercase."""
    n = name
    # Strip leading buy-in like "$1,500 " or "$200 "
    n = re.sub(r"^\$[\d,]+\s+", "", n)
    # Strip "- Day " patterns
    n = re.sub(r"\s*-\s*Day\s+", " ", n)
    # Normalize common game-name spellings
    n = (n.replace("No-Limit Hold'Em", "NLH")
         .replace("No Limit Hold'Em", "NLH")
         .replace("No-Limit Holdem", "NLH")
         .replace("Pot-Limit Omaha", "PLO")
         .replace("Pot Limit Omaha", "PLO")
         .replace("Omaha 8 or Better", "O/8")
         .replace("Omaha Hi-Lo", "O/8")
         .replace("Hold'em", "NLH"))
    # Lowercase + collapse whitespace + strip non-alphanumerics for the comparison key
    n = n.lower()
    n = re.sub(r"[^a-z0-9]+", " ", n).strip()
    n = re.sub(r"\s+", " ", n)
    return n


# wizardofviz venue → our slug (canonical mapping; everything else falls through to slug_for_venue_display)
_WOV_VENUE_TO_SLUG = {
    "Horseshoe/Paris (WSOP)": "wsop-paris",
    "Horseshoe (Circuit)": None,    # not in our venue set; skip
    "Caesars Palace": None,         # not in our venue set; skip
    "Venetian": "venetian",
    "Wynn": "wynn",
    "Aria": "aria",
    "MGM": "mgm-grand",
    "Orleans": "orleans",
    "South Point": "south-point",
    "Golden Nugget": "golden-nugget",
}


def _build_wov_index(wov_events: list[dict], venues: list[dict]) -> dict:
    """Index wizardofviz events by (slug, date_pt, normalized_event_name).
    Also stores by-buy-in-fallback key (slug, date_pt, buy_in, game) as a secondary lookup."""
    from datetime import date as _date
    primary: dict[tuple[str, _date, str], dict] = {}
    fallback: dict[tuple[str, _date, int, str], dict] = {}
    for e in wov_events:
        venue = e.get("Location", "")
        slug = _WOV_VENUE_TO_SLUG.get(venue)
        if slug is None:
            continue
        date_str = e.get("DATE", "")
        try:
            day = _date.fromisoformat(date_str)
        except (ValueError, TypeError):
            continue
        name = e.get("Event", "")
        if not name:
            continue
        norm = _normalize_event_name(name)
        primary.setdefault((slug, day, norm), e)
        buy_in = e.get("Buy-in")
        game = (e.get("Game") or "").strip().upper()
        if buy_in is not None and game:
            fallback.setdefault((slug, day, int(buy_in), game), e)
    return {"primary": primary, "fallback": fallback}


def _enrich_with_wov(t: Tournament, slug_lookup: dict, wov_index: dict) -> Tournament:
    slug = slug_lookup.get(t.venue_display)
    if not slug:
        return t
    norm = _normalize_event_name(t.event_name)
    wov = wov_index["primary"].get((slug, t.date_pt, norm))
    if wov is None:
        # Fallback by buy-in + game category
        if t.buy_in_usd is not None:
            game_key = t.game.strip().upper()
            wov = wov_index["fallback"].get((slug, t.date_pt, t.buy_in_usd, game_key))
    if wov is None:
        return t
    return dataclasses.replace(
        t,
        starting_stack=_safe_int(wov.get("Starting Stack")),
        level_minutes=_safe_str(wov.get("Blind Levels")),
        handed=_safe_int(wov.get("Handed")),
        rake_usd=_safe_float(wov.get("Rake")),
        rake_pct=_safe_float(wov.get("Rake %")),
    )


def _safe_int(v):
    try:
        return int(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _safe_float(v):
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _safe_str(v):
    s = str(v).strip() if v not in (None, "") else ""
    return s or None


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build vegas-poker JSON feed from XLSX")
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--venues", type=Path, required=True)
    parser.add_argument("--out-dir", type=Path, required=True)
    args = parser.parse_args(argv)
    build_feed(args.xlsx, args.venues, args.out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
