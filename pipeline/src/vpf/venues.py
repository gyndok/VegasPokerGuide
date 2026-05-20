from datetime import date as _date
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from typing import Optional
import yaml


def extract_schedule_hyperlinks(ws: Worksheet) -> list[str]:
    """Walk every cell of the Schedule tab and collect external hyperlink targets."""
    urls: list[str] = []
    seen: set[str] = set()
    for row in ws.iter_rows():
        for cell in row:
            link = cell.hyperlink
            if link and link.target and link.target.startswith("http"):
                if link.target not in seen:
                    seen.add(link.target)
                    urls.append(link.target)
    return urls


def _detect_venue_blocks(ws: Worksheet) -> dict[str, int]:
    """Find the venue header row (the row containing "Date" in column A, immediately
    followed by venue blocks). Returns {venue_display: start_col_index_0based}.

    The header row in the 2026 sheet is row 6 (column A == "Date"), and the venue
    names live on row 5 (one row above). We scan row 5 for non-empty cells that
    look like venue names.
    """
    # Find "Date" in column A; the venue names are on the prior row.
    date_row = None
    for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        if r_idx > 30:
            break  # don't scan the whole sheet
        a = row[0].value
        if isinstance(a, str) and a.strip().lower() == "date":
            date_row = r_idx
            break
    if not date_row:
        return {}

    venue_row_idx = date_row - 1
    venue_row = list(ws.iter_rows(min_row=venue_row_idx, max_row=venue_row_idx, values_only=False))[0]

    blocks: dict[str, int] = {}
    for col_idx, cell in enumerate(venue_row):
        v = cell.value
        if isinstance(v, str) and v.strip():
            # Strip the date-range suffix in parentheses, e.g.,
            # "Venetian (18/05 - 2/08)" -> "Venetian"
            name = v.split("(")[0].strip()
            if name:
                blocks[name] = col_idx
    return blocks


def extract_event_hyperlinks(ws: Worksheet) -> dict[tuple[str, _date, str], str]:
    """For each cell in the Schedule tab that is an EVENT cell (has an event name
    AND a hyperlink), return a mapping (venue_display, date, event_name) -> url.

    The Schedule tab is laid out as a date column (A) plus 8 venue blocks of 6
    columns each. Event names live in column index 4 of each venue block
    (zero-based within the block: Start, LR, Event, Buy-in, RE, Guarantee ->
    Event is index 2). Date is sticky from the most recent non-empty cell in
    column A.
    """
    from datetime import datetime as _dt

    venue_blocks = _detect_venue_blocks(ws)

    out: dict[tuple[str, _date, str], str] = {}
    current_date: _date | None = None
    for row in ws.iter_rows():
        # Update current_date from column A.
        a = row[0].value if row else None
        if isinstance(a, _dt):
            current_date = a.date()
        elif isinstance(a, _date):
            current_date = a
        if current_date is None:
            continue

        for venue_name, start_col in venue_blocks.items():
            # Event cell is at column offset +2 from venue start (0-indexed within block).
            event_col_idx = start_col + 2
            if event_col_idx >= len(row):
                continue
            cell = row[event_col_idx]
            link = cell.hyperlink
            if link is None or not link.target or not link.target.startswith("http"):
                continue
            event_name = str(cell.value or "").strip()
            if not event_name:
                continue
            key = (venue_name, current_date, event_name)
            # First-write wins so we preserve the earliest cell match.
            if key not in out:
                out[key] = link.target
    return out


def load_venues(yaml_path: Path) -> list[dict]:
    with open(yaml_path, "r") as f:
        return yaml.safe_load(f) or []


def merge_pdf_urls(venues: list[dict], discovered_urls: list[str]) -> list[dict]:
    """For each venue with an empty structure_pdf_url, fill in the first discovered URL
    whose host contains one of the venue's match_terms (case-insensitive).
    Curated values win — non-empty existing URLs are preserved.
    """
    result = []
    for v in venues:
        v = dict(v)  # copy
        if v.get("structure_pdf_url"):
            result.append(v)
            continue
        for url in discovered_urls:
            url_lower = url.lower()
            if any(term.lower() in url_lower for term in v.get("match_terms", [])):
                v["structure_pdf_url"] = url
                break
        result.append(v)
    return result


def slug_for_venue_display(display: str, venues: list[dict]) -> Optional[str]:
    """Match a sheet venue string to a curated slug via case-insensitive 'match_terms' lookup."""
    if not display:
        return None
    haystack = display.upper()
    for v in venues:
        for term in v.get("match_terms", []):
            if term.upper() in haystack:
                return v["slug"]
    return None
