from openpyxl import load_workbook
from vpf.tournaments import Tournament, parse_list_tab


def test_parses_first_known_row(sheet_xlsx):
    wb = load_workbook(sheet_xlsx, data_only=True)
    tournaments, warnings = parse_list_tab(wb["List"])
    # Row 2 in the fixture: Mon 18/5 Venetian NLH 1A $600 UL $150,000
    first = tournaments[0]
    assert first.venue_display == "Venetian"
    assert first.date_pt.isoformat() == "2026-05-18"
    assert first.start_at_pt.isoformat() == "2026-05-18T11:10:00-07:00"
    assert first.late_reg_close_at_pt.isoformat() == "2026-05-18T17:55:00-07:00"
    assert first.event_name == "NLH 1A"
    assert first.game == "NLH"
    assert first.game_category == "nlh"
    assert first.buy_in_usd == 600
    assert first.guarantee_usd == 150000
    assert first.re_entry.type == "unlimited"
    assert first.is_day2 is False


def test_detects_day2_rows(sheet_xlsx):
    wb = load_workbook(sheet_xlsx, data_only=True)
    tournaments, _ = parse_list_tab(wb["List"])
    assert any(t.is_day2 and "Day 2" in t.event_name for t in tournaments)


def test_total_row_count_in_expected_range(sheet_xlsx):
    wb = load_workbook(sheet_xlsx, data_only=True)
    tournaments, _ = parse_list_tab(wb["List"])
    assert 1000 < len(tournaments) < 2000


def test_flight_group_strips_suffix(sheet_xlsx):
    wb = load_workbook(sheet_xlsx, data_only=True)
    tournaments, _ = parse_list_tab(wb["List"])
    nlh_1a = next(t for t in tournaments if t.event_name == "NLH 1A" and t.venue_display == "Venetian")
    assert nlh_1a.flight_group == "NLH"


def test_no_guarantee_becomes_none(sheet_xlsx):
    # Day 2 rows have blank buy-in/guarantee — must be tolerated.
    wb = load_workbook(sheet_xlsx, data_only=True)
    tournaments, _ = parse_list_tab(wb["List"])
    day2 = next(t for t in tournaments if t.is_day2)
    assert day2.buy_in_usd is None or day2.guarantee_usd is None
