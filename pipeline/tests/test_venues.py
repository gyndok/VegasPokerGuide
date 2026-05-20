from openpyxl import load_workbook
from pathlib import Path
import pytest
from vpf.venues import extract_schedule_hyperlinks, extract_event_hyperlinks, load_venues, slug_for_venue_display


def test_extract_event_hyperlinks_returns_per_event_urls(sheet_xlsx):
    wb = load_workbook(sheet_xlsx)
    mapping = extract_event_hyperlinks(wb["Schedule 2026"])
    # At least a few hundred per-event links should be present.
    assert len(mapping) > 200
    # Venetian NLH 1A is a known event; check we have a URL for it.
    from datetime import date
    venetian_keys = [k for k in mapping if k[0] == "Venetian" and "NLH 1A" in k[2]]
    assert venetian_keys, "expected at least one Venetian NLH 1A entry"
    # Each Venetian entry should point to a venetianlasvegas.com URL.
    for k in venetian_keys[:3]:
        assert "venetianlasvegas.com" in mapping[k]


def test_extracts_pdf_urls(sheet_xlsx):
    wb = load_workbook(sheet_xlsx)  # NOT data_only — preserve hyperlinks
    urls = extract_schedule_hyperlinks(wb["Schedule 2026"])
    # urls is a set of absolute URLs
    assert any("venetianlasvegas.com" in u for u in urls)
    assert any("wynnresorts.com" in u for u in urls)
    assert any("wsop.gg-global-cdn.com" in u for u in urls)
    assert any("southpointcasino.com" in u for u in urls)


def test_urls_are_unique(sheet_xlsx):
    wb = load_workbook(sheet_xlsx)
    urls = extract_schedule_hyperlinks(wb["Schedule 2026"])
    assert len(urls) == len(set(urls))


VENUES_YML = Path(__file__).parent.parent / "venues.yml"


def test_load_venues_returns_eight():
    venues = load_venues(VENUES_YML)
    assert len(venues) == 8
    slugs = {v["slug"] for v in venues}
    assert {"wsop-paris", "wynn", "venetian", "aria", "mgm-grand", "orleans", "south-point", "golden-nugget"} == slugs


def test_merge_pdf_urls_fills_empty_only():
    from vpf.venues import merge_pdf_urls
    venues = [
        {"slug": "wynn", "structure_pdf_url": "https://existing.example/wynn.pdf", "match_terms": ["Wynn"]},
        {"slug": "venetian", "structure_pdf_url": "", "match_terms": ["Venetian"]},
        {"slug": "ghost", "structure_pdf_url": "", "match_terms": ["Ghost"]},
    ]
    discovered = [
        "https://cdn.wynnresorts.com/late.pdf",
        "https://venetianlasvegas.com/structure.pdf",
    ]
    merged = merge_pdf_urls(venues, discovered)
    assert merged[0]["structure_pdf_url"] == "https://existing.example/wynn.pdf"  # curated wins
    assert merged[1]["structure_pdf_url"] == "https://venetianlasvegas.com/structure.pdf"
    assert merged[2]["structure_pdf_url"] == ""  # no match left empty


@pytest.mark.parametrize("display,expected", [
    ("Venetian", "venetian"),
    ("Wynn", "wynn"),
    ("WSOP Paris/Horseshoe", "wsop-paris"),
    ("Horseshoe", "wsop-paris"),
    ("Aria / PokerGo", "aria"),
    ("Orleans", "orleans"),
    ("South Point", "south-point"),
    ("Golden Nugget", "golden-nugget"),
    ("MGM Grand", "mgm-grand"),
    ("Unknown Venue", None),
])
def test_slug_for_venue_display(display, expected):
    venues = load_venues(VENUES_YML)
    assert slug_for_venue_display(display, venues) == expected
