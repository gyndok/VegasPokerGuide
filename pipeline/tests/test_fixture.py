from openpyxl import load_workbook


def test_fixture_loads(sheet_xlsx):
    wb = load_workbook(sheet_xlsx, data_only=True)
    assert wb.sheetnames == ["Schedule 2026", "Buy-ins 2026", "List"]
    assert wb["List"].max_row > 1000  # ~1300 rows expected
